"""Safe master-tracker spreadsheet parsing and normalization.

Safety properties, each corresponding to a real hazard in migrating a
hand-maintained tracker:

* **Formulas are data.** The workbook is read with ``data_only=False`` so a cell
  containing ``=SUM(...)`` yields the literal formula string, which is recorded
  verbatim and flagged — never evaluated, and never silently turned into a number.
* **Macros are refused.** ``.xlsm``/``.xlsb``/legacy ``.xls`` are rejected before
  any parsing happens.
* **Nothing is guessed.** Ambiguous dates, unknown statuses, and unresolvable
  entities become typed row errors, not best-effort values.
* **Every row keeps provenance.** The original cells and a fingerprint of the
  business key are preserved so a disputed value traces back to its source cell.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from app.imports.enums import (
    REJECTED_EXTENSIONS,
    REQUIRED_TRACKER_COLUMNS,
    ImportErrorCode,
    SpreadsheetFormat,
    TrackerColumn,
)
from app.licensing.enums import FilingChannel, LicenseStatus

MAX_CELL_LENGTH = 2000
#: Formula-looking cell prefixes. Recorded as text and flagged for review.
_FORMULA_PREFIXES = ("=", "+@", "@")


class TrackerParseError(Exception):
    """The spreadsheet could not be parsed at all."""


@dataclass(slots=True)
class RawRow:
    """One spreadsheet row, exactly as read."""

    row_number: int
    cells: dict[str, str]
    formula_cells: list[str] = field(default_factory=list)


@dataclass(slots=True)
class SheetPreview:
    """What an Admin sees before choosing a column mapping."""

    sheet_names: list[str]
    selected_sheet: str | None
    headers: list[str]
    sample_rows: list[dict[str, str]]
    total_rows: int
    formula_cell_count: int
    notes: list[str] = field(default_factory=list)


@dataclass(slots=True)
class RowIssue:
    code: str
    message: str
    column: str | None = None


@dataclass(slots=True)
class NormalizedRow:
    """A row translated into canonical fields, with any problems attached."""

    row_number: int
    source: dict[str, str]
    values: dict[str, Any] = field(default_factory=dict)
    issues: list[RowIssue] = field(default_factory=list)
    fingerprint: str = ""

    @property
    def has_errors(self) -> bool:
        return bool(self.issues)


def reject_unsafe_extension(filename: str) -> None:
    """Refuse macro-enabled and legacy binary workbooks before parsing."""
    lowered = (filename or "").lower()
    for extension in REJECTED_EXTENSIONS:
        if lowered.endswith(extension):
            raise TrackerParseError(
                f"{extension} workbooks are not accepted because they can carry macros "
                "or require legacy binary parsing. Save the tracker as .xlsx or .csv."
            )


def detect_format(filename: str) -> str:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return SpreadsheetFormat.CSV.value
    if lowered.endswith(".xlsx") or lowered.endswith(".xltx"):
        return SpreadsheetFormat.XLSX.value
    raise TrackerParseError(f"Unsupported tracker file type: {filename!r}")


def _clean_cell(value: Any) -> tuple[str, bool]:
    """Stringify a cell. Returns ``(text, is_formula)``.

    Dates keep ISO form so downstream parsing is unambiguous. Formula strings are
    preserved verbatim and reported, never computed.
    """
    if value is None:
        return "", False
    if isinstance(value, datetime):
        return value.date().isoformat(), False
    if isinstance(value, date):
        return value.isoformat(), False
    if isinstance(value, bool):
        return ("true" if value else "false"), False
    if isinstance(value, int | float):
        # Avoid "1234.0" for whole numbers read from a numeric cell.
        text = str(int(value)) if float(value).is_integer() else str(value)
        return text, False
    text = str(value).strip()
    is_formula = text.startswith(_FORMULA_PREFIXES) and len(text) > 1
    return text[:MAX_CELL_LENGTH], is_formula


def read_xlsx(
    content: bytes, *, sheet_name: str | None = None, max_rows: int = 20000
) -> tuple[list[str], list[RawRow], list[str], list[str]]:
    """Read an XLSX. Returns ``(headers, rows, sheet_names, notes)``."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is pinned
        raise TrackerParseError("openpyxl is required to read XLSX trackers.") from exc

    try:
        # data_only=False: never surface a cached formula result as if it were data.
        workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
    except Exception as exc:
        raise TrackerParseError(f"The workbook could not be opened: {type(exc).__name__}") from exc

    notes: list[str] = []
    try:
        sheet_names = list(workbook.sheetnames)
        if sheet_name and sheet_name not in sheet_names:
            raise TrackerParseError(
                f"Sheet {sheet_name!r} is not present. Available: {sheet_names}"
            )
        sheet = workbook[sheet_name] if sheet_name else workbook[sheet_names[0]]

        headers: list[str] = []
        rows: list[RawRow] = []
        for index, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
            if raw is None:
                continue
            if not headers:
                # First non-empty row is the header row.
                candidate = [_clean_cell(cell)[0] for cell in raw]
                if any(value for value in candidate):
                    headers = _dedupe_headers(candidate)
                continue
            if len(rows) >= max_rows:
                notes.append(f"Stopped after the {max_rows}-row limit.")
                break
            cells: dict[str, str] = {}
            formulas: list[str] = []
            empty = True
            for position, header in enumerate(headers):
                value = raw[position] if position < len(raw) else None
                text, is_formula = _clean_cell(value)
                cells[header] = text
                if text:
                    empty = False
                if is_formula:
                    formulas.append(header)
            if empty:
                continue
            rows.append(RawRow(row_number=index, cells=cells, formula_cells=formulas))
    finally:
        workbook.close()

    if not headers:
        raise TrackerParseError("No header row was found in the selected sheet.")
    return headers, rows, sheet_names, notes


def read_csv(content: bytes, *, max_rows: int = 20000) -> tuple[list[str], list[RawRow], list[str]]:
    """Read a CSV tracker, tolerating a UTF-8 BOM."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise TrackerParseError("The CSV is not valid UTF-8 or Windows-1252.") from exc

    reader = csv.reader(io.StringIO(text))
    notes: list[str] = []
    headers: list[str] = []
    rows: list[RawRow] = []
    for index, raw in enumerate(reader, start=1):
        if not raw or not any(cell.strip() for cell in raw):
            continue
        if not headers:
            headers = _dedupe_headers([cell.strip() for cell in raw])
            continue
        if len(rows) >= max_rows:
            notes.append(f"Stopped after the {max_rows}-row limit.")
            break
        cells: dict[str, str] = {}
        formulas: list[str] = []
        for position, header in enumerate(headers):
            value = raw[position] if position < len(raw) else ""
            cleaned, is_formula = _clean_cell(value)
            cells[header] = cleaned
            if is_formula:
                formulas.append(header)
        rows.append(RawRow(row_number=index, cells=cells, formula_cells=formulas))

    if not headers:
        raise TrackerParseError("No header row was found in the CSV.")
    return headers, rows, notes


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Ensure header names are unique and non-empty so mapping is unambiguous."""
    seen: dict[str, int] = {}
    result: list[str] = []
    for position, header in enumerate(headers, start=1):
        name = (header or "").strip() or f"column_{position}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        result.append(name)
    return result


def preview(
    content: bytes, *, filename: str, sheet_name: str | None = None, sample_size: int = 5
) -> SheetPreview:
    """Parse enough of the file for an Admin to confirm sheet and mapping."""
    reject_unsafe_extension(filename)
    fmt = detect_format(filename)
    if fmt == SpreadsheetFormat.XLSX.value:
        headers, rows, sheet_names, notes = read_xlsx(content, sheet_name=sheet_name)
        selected = sheet_name or (sheet_names[0] if sheet_names else None)
    else:
        headers, rows, notes = read_csv(content)
        sheet_names, selected = [], None

    formula_cells = sum(len(row.formula_cells) for row in rows)
    if formula_cells:
        notes.append(
            f"{formula_cells} cell(s) contain formulas. They are imported as literal "
            "text and never evaluated; review them before applying."
        )
    return SheetPreview(
        sheet_names=sheet_names,
        selected_sheet=selected,
        headers=headers,
        sample_rows=[row.cells for row in rows[:sample_size]],
        total_rows=len(rows),
        formula_cell_count=formula_cells,
        notes=notes,
    )


# --------------------------------------------------------------------------
# Normalization
# --------------------------------------------------------------------------

#: Tracker status vocabulary observed in hand-maintained spreadsheets, mapped onto
#: the canonical licence status enum. Anything unmatched becomes a typed error
#: rather than a guess.
STATUS_SYNONYMS: dict[str, str] = {
    "active": LicenseStatus.ACTIVE.value,
    "current": LicenseStatus.ACTIVE.value,
    "licensed": LicenseStatus.ACTIVE.value,
    "approved": LicenseStatus.APPROVED.value,
    "issued": LicenseStatus.ACTIVE.value,
    "pending": LicenseStatus.PENDING_REGULATOR.value,
    "pending regulator": LicenseStatus.PENDING_REGULATOR.value,
    "submitted": LicenseStatus.PENDING_REGULATOR.value,
    "in progress": LicenseStatus.APPLICATION_IN_PROGRESS.value,
    "application in progress": LicenseStatus.APPLICATION_IN_PROGRESS.value,
    "applying": LicenseStatus.APPLICATION_IN_PROGRESS.value,
    "renewal in progress": LicenseStatus.RENEWAL_IN_PROGRESS.value,
    "renewing": LicenseStatus.RENEWAL_IN_PROGRESS.value,
    "expired": LicenseStatus.EXPIRED.value,
    "lapsed": LicenseStatus.EXPIRED.value,
    "reinstating": LicenseStatus.REINSTATING.value,
    "suspended": LicenseStatus.SUSPENDED.value,
    "revoked": LicenseStatus.REVOKED.value,
    "surrendered": LicenseStatus.SURRENDERED.value,
    "withdrawn": LicenseStatus.SURRENDERED.value,
    "closed": LicenseStatus.SURRENDERED.value,
    "not required": LicenseStatus.NOT_REQUIRED.value,
    "n/a": LicenseStatus.NOT_REQUIRED.value,
    "na": LicenseStatus.NOT_REQUIRED.value,
    "exempt": LicenseStatus.NOT_REQUIRED.value,
    "not started": LicenseStatus.NOT_STARTED.value,
    "unknown": LicenseStatus.UNKNOWN.value,
}

CHANNEL_SYNONYMS: dict[str, str] = {
    "nmls": FilingChannel.NMLS.value,
    "yes": FilingChannel.NMLS.value,
    "nmls yes": FilingChannel.NMLS.value,
    "state portal": FilingChannel.STATE_PORTAL.value,
    "state": FilingChannel.STATE_PORTAL.value,
    "online": FilingChannel.STATE_PORTAL.value,
    "portal": FilingChannel.STATE_PORTAL.value,
    "local portal": FilingChannel.LOCAL_PORTAL.value,
    "city": FilingChannel.LOCAL_PORTAL.value,
    "paper": FilingChannel.PAPER.value,
    "mail": FilingChannel.PAPER.value,
    "hard copy": FilingChannel.PAPER.value,
    "email": FilingChannel.EMAIL.value,
    "vendor": FilingChannel.VENDOR_MANAGED.value,
    "vendor managed": FilingChannel.VENDOR_MANAGED.value,
    "multiple": FilingChannel.MULTIPLE_CHANNELS.value,
    "no": FilingChannel.UNKNOWN.value,
    "n/a": FilingChannel.UNKNOWN.value,
    "unknown": FilingChannel.UNKNOWN.value,
}

#: Accepted date layouts. Deliberately excludes ambiguous two-digit years and
#: day-first forms: "03/04/25" cannot be disambiguated, so it is an error.
_DATE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d %b %Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%Y/%m/%d",
)
_AMBIGUOUS_DATE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2}$")


def parse_date(value: str) -> tuple[date | None, str | None]:
    """Parse a tracker date. Returns ``(date, error_message)``."""
    text = (value or "").strip()
    if not text:
        return None, None
    if text.lower() in ("n/a", "na", "none", "tbd", "-", "unknown"):
        return None, None
    if _AMBIGUOUS_DATE.match(text):
        return None, (f"{text!r} uses a two-digit year, which is ambiguous. Use YYYY-MM-DD.")
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    # Excel serial dates occasionally survive as bare integers.
    if text.isdigit() and 20000 <= int(text) <= 60000:
        from datetime import timedelta

        return date(1899, 12, 30) + timedelta(days=int(text)), None
    return None, f"{text!r} is not a recognised date. Use YYYY-MM-DD."


def parse_amount(value: str) -> tuple[float | None, str | None]:
    text = (value or "").strip()
    if not text or text.lower() in ("n/a", "na", "none", "-"):
        return None, None
    cleaned = re.sub(r"[$,\s]", "", text)
    try:
        amount = float(cleaned)
    except ValueError:
        return None, f"{text!r} is not a recognised amount."
    if amount < 0:
        return None, f"{text!r} is negative."
    return amount, None


def normalize_status(value: str) -> tuple[str | None, str | None]:
    text = (value or "").strip()
    if not text:
        return LicenseStatus.UNKNOWN.value, None
    key = re.sub(r"\s+", " ", text.lower())
    if key in STATUS_SYNONYMS:
        return STATUS_SYNONYMS[key], None
    upper = text.upper().replace(" ", "_")
    if upper in {member.value for member in LicenseStatus}:
        return upper, None
    return None, f"{text!r} is not a recognised licence status."


def normalize_channel(value: str) -> tuple[str, str | None]:
    text = (value or "").strip()
    if not text:
        return FilingChannel.UNKNOWN.value, None
    key = re.sub(r"\s+", " ", text.lower())
    if key in CHANNEL_SYNONYMS:
        return CHANNEL_SYNONYMS[key], None
    upper = text.upper().replace(" ", "_")
    if upper in {member.value for member in FilingChannel}:
        return upper, None
    # Unknown channel is recorded as UNKNOWN with a note: a filing channel we
    # cannot classify must not silently become "NMLS".
    return FilingChannel.UNKNOWN.value, f"{text!r} is not a recognised filing channel."


def row_fingerprint(values: dict[str, Any]) -> str:
    """Digest of the business key, used to detect repeat imports."""
    key = "|".join(
        str(values.get(column) or "").strip().lower()
        for column in (
            TrackerColumn.LEGAL_ENTITY.value,
            TrackerColumn.JURISDICTION.value,
            TrackerColumn.LICENSE_TYPE.value,
            TrackerColumn.LICENSE_NUMBER.value,
        )
    )
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


def validate_mapping(mapping: dict[str, str], headers: list[str]) -> list[str]:
    """Validate a column mapping before any row is touched."""
    problems: list[str] = []
    valid_targets = {member.value for member in TrackerColumn}
    for target, source in mapping.items():
        if target not in valid_targets:
            problems.append(f"Unknown destination field {target!r}.")
        elif source and source not in headers:
            problems.append(f"Column {source!r} for {target} is not present in the file.")
    for required in REQUIRED_TRACKER_COLUMNS:
        if not mapping.get(required):
            problems.append(f"{required} must be mapped before the import can run.")
    return problems


def normalize_row(row: RawRow, mapping: dict[str, str]) -> NormalizedRow:
    """Translate one raw row into canonical fields, collecting typed issues."""
    result = NormalizedRow(row_number=row.row_number, source=dict(row.cells))

    def raw(target: str) -> str:
        column = mapping.get(target)
        return (row.cells.get(column, "") if column else "").strip()

    for target in (
        TrackerColumn.LEGAL_ENTITY.value,
        TrackerColumn.JURISDICTION.value,
        TrackerColumn.LICENSE_TYPE.value,
        TrackerColumn.LICENSE_NUMBER.value,
        TrackerColumn.NMLS_LICENSE_ID.value,
        TrackerColumn.VENDOR.value,
        TrackerColumn.REGULATOR.value,
        TrackerColumn.BOND_PROVIDER.value,
        TrackerColumn.BOND_NUMBER.value,
        TrackerColumn.RESPONSIBLE_OWNER.value,
        TrackerColumn.NOTES.value,
        TrackerColumn.NEXT_ACTION.value,
        TrackerColumn.SOURCE_DOCUMENT_REFERENCE.value,
    ):
        value = raw(target)
        if value:
            result.values[target] = value

    for required in REQUIRED_TRACKER_COLUMNS:
        if not result.values.get(required):
            code = {
                TrackerColumn.LEGAL_ENTITY.value: ImportErrorCode.UNRESOLVED_LEGAL_ENTITY.value,
                TrackerColumn.JURISDICTION.value: ImportErrorCode.UNRESOLVED_JURISDICTION.value,
                TrackerColumn.LICENSE_TYPE.value: ImportErrorCode.UNRESOLVED_LICENSE_TYPE.value,
            }[required]
            result.issues.append(
                RowIssue(code, f"{required} is empty and cannot be resolved.", required)
            )

    status, status_error = normalize_status(raw(TrackerColumn.STATUS.value))
    if status_error:
        result.issues.append(
            RowIssue(ImportErrorCode.INVALID_STATUS.value, status_error, TrackerColumn.STATUS.value)
        )
    else:
        result.values[TrackerColumn.STATUS.value] = status

    channel, channel_note = normalize_channel(raw(TrackerColumn.FILING_CHANNEL.value))
    result.values[TrackerColumn.FILING_CHANNEL.value] = channel
    if channel_note:
        result.issues.append(
            RowIssue(
                ImportErrorCode.INVALID_FILING_CHANNEL.value,
                channel_note,
                TrackerColumn.FILING_CHANNEL.value,
            )
        )

    date_columns = (
        TrackerColumn.ISSUE_DATE.value,
        TrackerColumn.EFFECTIVE_DATE.value,
        TrackerColumn.EXPIRATION_DATE.value,
        TrackerColumn.RENEWAL_DUE_DATE.value,
        TrackerColumn.BOND_EFFECTIVE_DATE.value,
        TrackerColumn.BOND_EXPIRATION_DATE.value,
        TrackerColumn.ANNUAL_REPORT_DATE.value,
    )
    for column in date_columns:
        parsed, error = parse_date(raw(column))
        if error:
            result.issues.append(RowIssue(ImportErrorCode.INVALID_DATE.value, error, column))
        elif parsed is not None:
            result.values[column] = parsed

    amount, amount_error = parse_amount(raw(TrackerColumn.BOND_AMOUNT.value))
    if amount_error:
        result.issues.append(
            RowIssue(
                ImportErrorCode.INVALID_BOND_DETAILS.value,
                amount_error,
                TrackerColumn.BOND_AMOUNT.value,
            )
        )
    elif amount is not None:
        result.values[TrackerColumn.BOND_AMOUNT.value] = amount

    issue = result.values.get(TrackerColumn.ISSUE_DATE.value)
    expiration = result.values.get(TrackerColumn.EXPIRATION_DATE.value)
    if issue and expiration and expiration < issue:
        result.issues.append(
            RowIssue(
                ImportErrorCode.INVALID_DATE_SEQUENCE.value,
                f"Expiration {expiration.isoformat()} precedes issue {issue.isoformat()}.",
                TrackerColumn.EXPIRATION_DATE.value,
            )
        )

    # A bond with a provider or amount but no number is incomplete.
    if (
        result.values.get(TrackerColumn.BOND_PROVIDER.value)
        or result.values.get(TrackerColumn.BOND_AMOUNT.value)
    ) and not result.values.get(TrackerColumn.BOND_NUMBER.value):
        result.issues.append(
            RowIssue(
                ImportErrorCode.INVALID_BOND_DETAILS.value,
                "Bond details are present but the bond number is missing.",
                TrackerColumn.BOND_NUMBER.value,
            )
        )

    if row.formula_cells:
        result.values["_formula_columns"] = list(row.formula_cells)

    result.fingerprint = row_fingerprint(result.values)
    return result


def normalize_rows(rows: list[RawRow], mapping: dict[str, str]) -> list[NormalizedRow]:
    """Normalize every row and flag duplicates *within the file*."""
    normalized = [normalize_row(row, mapping) for row in rows]
    seen: dict[str, int] = {}
    for row in normalized:
        if row.has_errors:
            continue
        if row.fingerprint in seen:
            row.issues.append(
                RowIssue(
                    ImportErrorCode.DUPLICATE_ROW_IN_FILE.value,
                    f"Duplicates row {seen[row.fingerprint]} in the same file.",
                )
            )
        else:
            seen[row.fingerprint] = row.row_number
    return normalized


__all__ = [
    "CHANNEL_SYNONYMS",
    "MAX_CELL_LENGTH",
    "STATUS_SYNONYMS",
    "NormalizedRow",
    "RawRow",
    "RowIssue",
    "SheetPreview",
    "TrackerParseError",
    "detect_format",
    "normalize_channel",
    "normalize_row",
    "normalize_rows",
    "normalize_status",
    "parse_amount",
    "parse_date",
    "preview",
    "read_csv",
    "read_xlsx",
    "reject_unsafe_extension",
    "row_fingerprint",
    "validate_mapping",
]
