"""Build the portal's minimal current-tracker snapshot.

Only the ``DB`` and ``Non Licensed States`` sheets are consulted. The generated
JSON deliberately omits licence, bond, and filing identifiers; the portal view
needs dates and operational context, not sensitive credential-like numbers.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import warnings
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DB_SHEET = "DB"
NON_LICENSED_SHEET = "Non Licensed States"
PLACEHOLDERS = {"", "-", "na", "n/a", "nl", "nr", "not required", "none"}
DATE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%b %d %Y",
    "%B %d %Y",
    "%b %d, %Y",
    "%B %d, %Y",
)

warnings.filterwarnings(
    "ignore",
    message=r"xl/pivotCache/_rels/.* contains invalid dependency definitions",
    category=UserWarning,
    module=r"openpyxl\.packaging\.relationship",
)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return None if text.casefold() in PLACEHOLDERS or set(text) == {"-"} else text


def _status_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return None if not text or set(text) == {"-"} else text


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if text is None or text.casefold() in {"perpetual", "tbd", "unknown"}:
        return None
    normalized = re.sub(r"\s*['\u2019]\s*", " ", text)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(normalized, date_format).date()
        except ValueError:
            continue
    return None


def _event(
    *,
    row_number: int,
    state: str,
    abbreviation: str | None,
    jurisdiction_type: str | None,
    tracker_status: str | None,
    item_type: str,
    item_name: str,
    due_value: Any,
    agency: Any = None,
    owner: Any = None,
    notes: Any = None,
    source_column: str,
) -> dict[str, Any] | None:
    due_date = _date(due_value)
    if due_date is None:
        return None
    return {
        "event_id": f"db-{row_number}-{item_type.casefold().replace(' ', '-')}",
        "state": state,
        "abbreviation": abbreviation,
        "jurisdiction_type": jurisdiction_type,
        "tracker_status": tracker_status,
        "item_type": item_type,
        "item_name": item_name,
        "due_date": due_date.isoformat(),
        "agency": _text(agency),
        "owner": _text(owner),
        "notes": _text(notes),
        "source_row": row_number,
        "source_cell": f"{DB_SHEET}!{source_column}{row_number}",
    }


def build_snapshot(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        input_path,
        data_only=False,
        read_only=False,
        keep_links=False,
    )
    try:
        for required in (DB_SHEET, NON_LICENSED_SHEET):
            if required not in workbook.sheetnames:
                raise ValueError(f"Required sheet {required!r} is missing.")

        db_sheet = workbook[DB_SHEET]
        non_licensed_sheet = workbook[NON_LICENSED_SHEET]
        headers = [cell.value for cell in db_sheet[1]]
        expected = {
            0: "State/UT",
            1: "St Abbr",
            2: "Status",
            3: "States or UT",
            4: "NMLS",
            6: "Renew Date",
            13: "Renew Date",
            18: "Due Date",
            24: "Other Document Renewal Date",
        }
        for index, heading in expected.items():
            if index >= len(headers) or headers[index] != heading:
                raise ValueError(
                    f"{DB_SHEET} column {index + 1} must be {heading!r}; "
                    f"found {headers[index] if index < len(headers) else None!r}."
                )

        events: list[dict[str, Any]] = []
        non_licensed: list[dict[str, Any]] = []
        populated_rows = 0
        jurisdictions: set[str] = set()

        for row_number, row in enumerate(
            db_sheet.iter_rows(min_row=2, max_col=27, values_only=True),
            start=2,
        ):
            state = _text(row[0])
            if state is None:
                continue
            populated_rows += 1
            jurisdictions.add(state.casefold())
            abbreviation = _text(row[1])
            tracker_status = _status_text(row[2])
            jurisdiction_type = _text(row[3])

            candidates = (
                _event(
                    row_number=row_number,
                    state=state,
                    abbreviation=abbreviation,
                    jurisdiction_type=jurisdiction_type,
                    tracker_status=tracker_status,
                    item_type="LICENSE",
                    item_name="License renewal",
                    due_value=row[6],
                    agency=row[7],
                    notes=row[10],
                    source_column="G",
                ),
                _event(
                    row_number=row_number,
                    state=state,
                    abbreviation=abbreviation,
                    jurisdiction_type=jurisdiction_type,
                    tracker_status=tracker_status,
                    item_type="BOND",
                    item_name="Bond renewal",
                    due_value=row[13],
                    agency=row[15],
                    notes=row[14],
                    source_column="N",
                ),
                _event(
                    row_number=row_number,
                    state=state,
                    abbreviation=abbreviation,
                    jurisdiction_type=jurisdiction_type,
                    tracker_status=tracker_status,
                    item_type="ANNUAL_REPORT",
                    item_name="Annual report",
                    due_value=row[18],
                    agency=row[20],
                    notes=row[19],
                    source_column="S",
                ),
                _event(
                    row_number=row_number,
                    state=state,
                    abbreviation=abbreviation,
                    jurisdiction_type=jurisdiction_type,
                    tracker_status=tracker_status,
                    item_type="OTHER_DOCUMENT",
                    item_name=_text(row[22]) or "Other document",
                    due_value=row[24],
                    owner=row[25],
                    notes=row[21],
                    source_column="Y",
                ),
            )
            events.extend(item for item in candidates if item is not None)

            if tracker_status and tracker_status.casefold() == "not licensed":
                non_licensed.append(
                    {
                        "record_id": f"db-{row_number}-not-licensed",
                        "state": state,
                        "abbreviation": abbreviation,
                        "jurisdiction_type": jurisdiction_type,
                        "nmls": _text(row[4]),
                        "reason": _text(row[8]),
                        "comments": _text(row[10]),
                        "source_row": row_number,
                    }
                )

        pivot_sources: list[str] = []
        for pivot in non_licensed_sheet._pivots:
            source = pivot.cache.cacheSource.worksheetSource
            if source and source.sheet and source.ref:
                pivot_sources.append(f"{source.sheet}!{source.ref}")

        events.sort(key=lambda item: (item["due_date"], item["state"], item["item_type"]))
        non_licensed.sort(key=lambda item: item["state"].casefold())
        source_bytes = input_path.read_bytes()
        stat = input_path.stat()
        return {
            "metadata": {
                "source_workbook": input_path.name,
                "source_sha256": hashlib.sha256(source_bytes).hexdigest(),
                "source_last_modified_at": datetime.fromtimestamp(
                    stat.st_mtime, tz=UTC
                ).isoformat(),
                "snapshot_generated_at": datetime.now(UTC).isoformat(),
                "source_sheets": [DB_SHEET, NON_LICENSED_SHEET],
                "db_rows": populated_rows,
                "tracked_jurisdictions": len(jurisdictions),
                "non_licensed_pivot_sources": pivot_sources,
                "data_minimization": (
                    "Licence, bond, annual-report, and other-document identifiers "
                    "are intentionally excluded from this portal snapshot."
                ),
            },
            "events": events,
            "non_licensed": non_licensed,
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    snapshot = build_snapshot(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(snapshot, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(args.output),
                "events": len(snapshot["events"]),
                "non_licensed": len(snapshot["non_licensed"]),
                "db_rows": snapshot["metadata"]["db_rows"],
            }
        )
    )


if __name__ == "__main__":
    main()
